# Прочитать сохранённый csv-файл из задания №17 и сохранить данные
# в excel-файл, кроме возраста – столбец с этими данными не нужен.
import csv
import openpyxl



with open("umba_tumba.csv", 'r', encoding="utf-8") as f:
    open_data = csv.reader(f)
    # for otkr in open_data:
    #     pprint(otkr)
    wb = openpyxl.Workbook()
    wb.create_sheet(title="Первый лист" , index=0)
    wb.remove(wb['Sheet'])
    print(wb.sheetnames)
    sheet = wb['Первый лист']
    print(sheet)
    for row_index, row in enumerate(open_data):
        for col_index, value in enumerate(row):
            cell = sheet.cell(row=row_index+1, column=col_index+1)
            cell.value = value
    wb.save("umba_tumba.xlsx")

import pandas as pd
wb = openpyxl.load_workbook("umba_tumba.xlsx")
wb = wb.active
wb_df = pd.DataFrame(wb)
wb_df = wb_df.drop(columns=["Возраст"])
wb.save("umba_tumba.xlsx")
