#Диплом - класс данных
import openpyxl


class Person_data():
    def __init__(self, name="", surname="", patronymic="", date_birth='', date_death='', gender=""):
        self.name = name
        self.date_brith = date_birth
        self.date_death = date_death
        self.gender = gender
        self.surname = surname
        self.patronymic = patronymic

    def load_person(self, name, surname, patronymic, date_birth, date_death, gender):
        bd_save = [(name, surname, patronymic, date_birth, date_death, gender)]
        zaglav = ["Имя", "Фамилия", "Отчество", "Дата Рождения", "Дата Смерти", "Пол"]
        wb = openpyxl.Workbook()
        wb.create_sheet(title="Первый лист", index=0)
        wb.remove(wb['Sheet'])
        sheet = wb['Первый лист']

        for col_index, value in enumerate(zaglav):
            cell = sheet.cell(row=1, column=col_index + 1)
            cell.value = value
        wb.save("Diplom.xlsx")

        wb_reading = openpyxl.load_workbook("Diplom.xlsx")
        sheet = wb_reading.active
        for row in bd_save:
            sheet.append(row)
        wb_reading.save("Diplom.xlsx")
        print("Готово!")

    def twoload_person(self, name_r, surname_r, patronymic_r, date_birth_r, date_death_r, gender_r):
        wb_reading = openpyxl.load_workbook("Diplom.xlsx")
        sheet = wb_reading.active
        load = [(name_r, surname_r, patronymic_r, date_birth_r, date_death_r, gender_r)]
        for row in load:
            sheet.append(row)
        wb_reading.save("Diplom.xlsx")
        print("Окей")

    def years_preson(self, name_year):
        wb_reading = openpyxl.load_workbook("Diplom.xlsx")
        sheet = wb_reading.active
        resultRow = 1
        for i in range(1, sheet.max_row + 1):
            value = sheet.cell(row=i, column=1).value
            if value == name_year:
                sheet.cell(row=resultRow, column=1).value = value
                for column in range(1, sheet.max_column):
                    date_birth = sheet[i][3].value
                    date_death = sheet[i][4].value
                    break


        print("Годы жизни")

    def unload_person(self, Search):
        wb_reading = openpyxl.load_workbook("Diplom.xlsx")
        sheet = wb_reading.active
        resultRow = 1
        for i in range(1, sheet.max_row + 1):
            value = sheet.cell(row=i, column=1).value
            if value == Search:
                sheet.cell(row=resultRow, column=1).value = value
                for column in range(1, sheet.max_column):
                    result = sheet[i][column].value
                    print(result)

    def reading(self):
        wb_reading = openpyxl.load_workbook("Diplom.xlsx", read_only=True)
        sheet = wb_reading.active
        for row in range(1, sheet.max_row+1):
            name = sheet[row][0].value
            surname = sheet[row][1].value
            patronymic = sheet[row][2].value
            date_birth = sheet[row][3].value
            date_death = sheet[row][4].value
            gender = sheet[row][5].value
            print(row, name, surname, patronymic, date_birth, date_death, gender)




